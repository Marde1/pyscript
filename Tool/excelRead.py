# -*- coding:utf-8 -*-

import openpyxl

class ExcelRead():

    def __init__(self,filename,sheetname):
        self.filename = filename
        self.wb = openpyxl.load_workbook(self.filename)
        self.sheet = self.wb[sheetname]
        self.maxrow_num = self.sheet.max_row
        self.maxcolumn_num = self.sheet.max_column

    def readData(self):

        # print("最大行：",self.maxrow_num)
        # print("最大列：",self.maxcolumn_num)
        data_dict = {}
        data_list =[]
        if(self.maxrow_num>0 and self.maxcolumn_num>0):
            for i in range(2,self.maxrow_num+1): #行
                for j in range(1,self.maxcolumn_num+1):#列
                    # print(i,j)
                    # print(self.sheet.cell(1,j).value)
                    # print(self.sheet.cell(i,j).value)
                    data_dict[self.sheet.cell(1,j).value] = self.sheet.cell(i,j).value
                # print(data_dict)
                data_list.append(data_dict)
                data_dict = {} #数据存入列表后清空字典，以便后面的数据存入
        return data_list

    def updateData(self,rownum,columnnum,value_str):
        """
        向excel中修改数据
        :param rownum: 行
        :param columnnum:列
        :param value_str: 需要写入的字符串
        :return:
        """
        self.sheet.cell(rownum,columnnum).value = value_str
        self.wb.save(self.filename)

    def updateResultData(self,rownum,value_str):
        """
        向excel中修改excle的最后一列的数据数据
        :param rownum: 行
        :param value_str: 需要写入的字符串
        :return:
        """
        self.updateData(rownum,self.maxcolumn_num,value_str)

if __name__ == "__main__":
    excel = ExcelRead(r"..\WebTours\webTour_data.xlsx","data")
    list01 = excel.readData()
    print(list01)